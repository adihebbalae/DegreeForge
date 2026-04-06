"""Convert all PDFs in scraped_data_corpus/ to .txt files and Excel to .txt.
Outputs go to scraped_data_corpus/txt/ for agent ingestion."""

import os
import sys

import fitz  # PyMuPDF
import openpyxl

CORPUS_DIR = os.path.join(os.path.dirname(__file__), "scraped_data_corpus")
OUTPUT_DIR = os.path.join(CORPUS_DIR, "txt")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def pdf_to_text(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    pages = []
    for page_num, page in enumerate(doc, 1):
        text = page.get_text("text")
        pages.append(f"--- PAGE {page_num} ---\n{text}")
    doc.close()
    return "\n\n".join(pages)

def xlsx_to_text(xlsx_path: str) -> str:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    output = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"=== SHEET: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            output.append("\t".join(cells))
    return "\n".join(output)

converted = 0
errors = []

for fname in sorted(os.listdir(CORPUS_DIR)):
    fpath = os.path.join(CORPUS_DIR, fname)
    if not os.path.isfile(fpath):
        continue

    base, ext = os.path.splitext(fname)
    ext = ext.lower()
    out_path = os.path.join(OUTPUT_DIR, f"{base}.txt")

    try:
        if ext == ".pdf":
            text = pdf_to_text(fpath)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"  OK  {fname} -> txt/{base}.txt  ({len(text):,} chars)")
            converted += 1
        elif ext == ".xlsx":
            text = xlsx_to_text(fpath)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"  OK  {fname} -> txt/{base}.txt  ({len(text):,} chars)")
            converted += 1
    except Exception as e:
        errors.append((fname, str(e)))
        print(f"  ERR {fname}: {e}", file=sys.stderr)

print(f"\nConverted: {converted}  |  Errors: {len(errors)}")
if errors:
    for name, err in errors:
        print(f"  FAILED: {name} — {err}")
